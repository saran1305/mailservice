# import pandas as pd
# import numpy as np
# from datetime import datetime, timedelta
# import json
# import logging
# import streamlit as st
# from openpyxl import load_workbook
# import warnings
# import requests
# from requests_oauthlib import OAuth2Session
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from urllib.parse import urlparse, parse_qs
# from dotenv import load_dotenv
# import os

# global client_id, tenant_id, client_secret, redirect_uri, authorization_base_url, token_endpoint, scope, app, send_mail_endpoint

# with open('config.json') as f:
#     app = json.load(f)

# # Load environment variables from .env file
# load_dotenv()

# warnings.simplefilter(action='ignore', category=UserWarning)

# client_id = os.getenv("CLIENT_ID")
# tenant_id = os.getenv("TENANT_ID")
# client_secret = os.getenv("CLIENT_SECRET")
# redirect_uri = os.getenv("REDIRECT_URI")
# authorization_base_url = os.getenv("AUTHORIZATION_BASE_URL")
# token_endpoint = os.getenv("TOKEN_ENDPOINT")
# scope = os.getenv("SCOPE")
# send_mail_endpoint = os.getenv("send_mail_endpoint")

# # Set up logging
# logging.basicConfig(level=logging.WARNING)

# def get_sheet_names(file_path):
#     wb = load_workbook(filename=file_path)
#     return wb.sheetnames

# def get_access_token():
#     # oauth = OAuth2Session(client_id, redirect_uri=redirect_uri, scope=scope)
#     # authorization_url, _ = oauth.authorization_url(authorization_base_url)
#     driver = webdriver.Chrome()
#     auth_url = "https://login.microsoftonline.com/db01513b-9352-48dc-9232-8fc9f4e6979f/oauth2/v2.0/authorize?client_id=dcc0356c-a07e-4272-9b1c-cb635e0fc2a0&response_type=id_token&redirect_uri=http%3A%2F%2Flocalhost:8501/&response_mode=form_post&scope=openid+email+profile"
#     driver.get(auth_url)
#     cookies = driver.get_cookies()
#     # print(cookies)

#     redirected_url = WebDriverWait(driver, 60).until(
#         EC.url_contains("localhost:8501/")
#     )
#     redirected_url = driver.current_url
#     driver.quit()
#     parsed_url = urlparse(redirected_url)
    

#     query_params = parse_qs(parsed_url.query)
#     auth_code = query_params.get('code', [''])[0]

#     token_request_params = {
#         'client_id': client_id,
#         'client_secret': client_secret,
#         'response_type': 'id_token',
#         'code': auth_code,
#         'redirect_uri': redirect_uri,
#         'grant_type': 'authorization_code',
#         'scope': scope
#     }
#     token_response = requests.post(token_endpoint, data=token_request_params)
#     token_data = token_response.json()
#     print(token_data)
#     expires_in = token_data.get('expires_in')
#     expiration_time = datetime.now() + timedelta(seconds=expires_in)
#     refresh_token = token_data.get('refresh_token')
#     return token_data.get('access_token'), expiration_time, refresh_token

# def refresh_access_token(refresh_token):
#     token_request_params = {
#         'client_id': client_id,
#         'client_secret': client_secret,
#         'refresh_token': refresh_token,
#         'grant_type': 'refresh_token',
#         'scope': scope
#     }
#     token_response = requests.post(token_endpoint, data=token_request_params)
#     token_data = token_response.json()
#     expires_in = token_data.get('expires_in')
#     expiration_time = datetime.now() + timedelta(seconds=expires_in)
#     return token_data.get('access_token'), expiration_time

# def mail_auth(access_token, From, to, cc, Subject, email_body):
#     headers = {
#         'Authorization': 'Bearer ' + access_token,
#         'Content-Type': 'application/json'
#     }
#     to_recipients = to.split(',')

#     to_recipient_list = [{"emailAddress": {"address": email.strip()}} for email in to_recipients]
#     cc_recipients = []
#     if cc and not pd.isna(cc):
#         cc_recipients_list = cc.split(',')
#         cc_recipients = [{"emailAddress": {"address": email.strip()}} for email in cc_recipients_list]
    
#     request_body = {
#         "message": {
#             "subject": f"{Subject}",
#             "body": {
#                 "contentType": "HTML",
#                 "content": f"{email_body}"
#             },
#             "toRecipients": to_recipient_list,
#             "ccRecipients": cc_recipients
#         },
#         "saveToSentItems": "true"
#     }
#     response = requests.post(send_mail_endpoint, headers=headers, json=request_body)
#     if response.status_code == 202:
#         return 'Email sent successfully!'
#     else:
#         return f'Email sending failed with status code: {response.status_code}'

# def send_email(access_token, receiver_emails, cc_emails, group, start_date, end_date, client_mapping):
#     sender_email = 'your_email@example.com'
#     subject = f"Weekly Schedule - {start_date.strftime('%Y-%b-%d')} to {end_date.strftime('%Y-%b-%d')} - Client Mapping: {client_mapping}"

#     From = sender_email
#     Subject = subject

#     header_style = '''
#     <style>
#         th {
#             background-color: #236418;
#             color: white;
#             font-weight: bold;
#         }
#     </style>
#     '''

#     email_body = f"""<p>Hello Team,</p>
#                     <p><b>{app['email_body']}</b></p>
#                     <p><b>{app['email_body_1']}</b></p>
#                     <p><b>{app['email_body_2']}</b></p>
#                     <p><b>{app['email_body_3']}</b></p>                    
#                     <p><b>Here is the schedule for the week {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')} for Client Mapping: {client_mapping}<b></p>
#                     <p><b>Important: This is an Automated mail, if any mistakes or doubts please, approach HR</b></p>
#                     {header_style}
#                     {group.drop(['Project head', 'Email Id', 'Client Mapping'], axis=1).rename(columns=lambda x: x.strftime('%Y-%b-%d') if isinstance(x, datetime) else x).to_html(index=False)}
#                     <p><b>Best regards,<br>Ideassion</b></p>
#                     """
    
#     try:
#         mail_auth(access_token, From, receiver_emails, cc_emails, Subject, email_body)
#         print("Email sent successfully.")
#     except Exception as e:
#         logging.warning(f"An error occurred while sending the email: {str(e)}")
#         raise Exception(f"An error occurred while sending the email: {str(e)}")

# def main():
#     st.image('ticjdbcnfxi1xrkyrkol.png', width=300)
#     st.title("Automatic Weekly Scheduler")
#     try:
#         today = datetime.today()
#         start_of_week = today + timedelta(days=(7 - today.weekday()))
#         end_of_week = start_of_week + timedelta(days= 6)
#         st.write(f"Automatically scheduling for the week {start_of_week.strftime('%Y-%b-%d')} to {end_of_week.strftime('%Y-%b-%d')}.")
#     except Exception as e:
#         logging.warning('The date series not available')

#     file_path = st.file_uploader("Upload Excel file", type=["xlsx"])
#     if file_path is not None:
#         try:
#             df = pd.read_excel(file_path)
#         except Exception as e:
#             logging.error(f"An error occurred while reading the Excel file: {str(e)}")
#             return

#         desired_sheet_names = get_sheet_names(file_path)
#         concatenated_schedules = []
#         multiple_sheets = []

#         for i, sheet_name in enumerate(desired_sheet_names):
#             try:
#                 df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=False)
#                 df['Emp.ID'] = df['Emp.ID'].fillna(0).astype(int)
#                 for column in df.columns:
#                     if column != 'Emp.ID':
#                         df[column] = df[column].map(lambda x: '' if isinstance(x, int) else x)

#                 date_columns = [col for col in df.columns if isinstance(col, datetime) and
#                                 col.date() >= start_of_week.date() and col.date() <= end_of_week.date() and col.weekday() not in [5, 6]]

#                 if date_columns:
#                     multiple_sheets.append(i)

#                     relevant_columns = ['Emp.ID','Name', 'Project head', 'Email Id', 'Client Mapping']
#                     if i == 0:
#                         date_schedule = df[relevant_columns + date_columns]
#                     elif i + 1:
#                         if i - 1 in multiple_sheets:
#                             date_schedule = df[date_columns]
#                         else:
#                             date_schedule = df[relevant_columns + date_columns]

#                     concatenated_schedules.append(date_schedule)

#                     current_sheet_data = df[relevant_columns + date_columns]

#                     if concatenated_schedules:
#                         concatenated_schedules[0].loc[:, 'Name'] = concatenated_schedules[0]['Name'].combine_first(
#                             current_sheet_data['Name'])
#                         concatenated_schedules[0].loc[:, 'Emp.ID'] = concatenated_schedules[0]['Emp.ID'].combine_first(
#                             current_sheet_data['Emp.ID'])
#                         concatenated_schedules[0].loc[:, 'Project head'] = concatenated_schedules[0]['Project head'].combine_first(
#                             current_sheet_data['Project head'])
#                         concatenated_schedules[0].loc[:, 'Email Id'] = concatenated_schedules[0]['Email Id'].combine_first(
#                             current_sheet_data['Email Id'])
#                         concatenated_schedules[0].loc[:, 'Client Mapping'] = concatenated_schedules[0]['Client Mapping'].combine_first(
#                             current_sheet_data['Client Mapping'])
#                     else:
#                         concatenated_schedules.append(current_sheet_data)
#             except Exception as e:
#                 logging.warning(f"An error occurred while processing sheet '{sheet_name}': {str(e)}")
#                 st.write(f"An error occurred while processing sheet '{sheet_name}': {str(e)}")

#         if multiple_sheets:
#             concatenated_schedule = pd.concat(concatenated_schedules, axis=1)

#         if concatenated_schedules:
#             concatenated_schedule = pd.concat(concatenated_schedules, axis=1)

#             if 'Client Mapping' in concatenated_schedule.columns:
#                 grouped_schedule = concatenated_schedule.groupby('Client Mapping', as_index=False)      
#                 send_email_button = st.button("Send Email")
#                 if send_email_button:
#                     access_token, expiration_time, refresh_token = get_access_token()

#                     if datetime.now() > expiration_time:
#                         print("Token has been expired")
#                         access_token, expiration_time = refresh_access_token(refresh_token)

#                     for client_mapping, group in grouped_schedule:
#                         group['Project head'] = group['Project head'].replace('nan', np.NaN)
#                         group = group.replace(np.nan, '', regex=True)

#                         project_lead_emails = []
#                         if 'arun_marx_info' in app:
#                             arun_marx_names = [name.strip().lower() for name in app['arun_marx_info']['Name'].split(',')]
#                             project_heads_names = set([name.strip() for head in group['Project head'] for name in head.split(',')])
#                             for project_head in project_heads_names:
#                                 if isinstance(project_head, str) and project_head.strip() != '':
#                                     project_head_lower = project_head.lower()
#                                     if any(name in project_head_lower for name in arun_marx_names):
#                                         project_lead_emails.append(app['arun_marx_info']['Email'])
#                                     else:
#                                         project_lead_email = df[df['Name'].str.lower() == project_head_lower]['Email Id'].unique()
#                                         project_lead_emails.extend(project_lead_email)
#                                 project_lead_emails = list(set(project_lead_emails))
#                         to = ', '.join(map(str, group['Email Id'].unique()))
#                         cc = ', '.join(map(str, project_lead_emails))
#                         send_email(access_token, to, cc, group, start_of_week, end_of_week, client_mapping)
#                         st.write(f"Email sent for Client Mapping: {client_mapping}")

#                 for client_mapping, group in grouped_schedule:
#                     st.subheader(f"Client Mapping: {client_mapping}")
#                     group['Project head'] = group['Project head'].replace('nan', np.NaN)
#                     group = group.replace(np.nan, '', regex=True)

#                     project_heads_names = set([name.strip() for head in group['Project head'] for name in head.split(',')])

#                     project_lead_emails = []
#                     if 'arun_marx_info' in app:
#                         arun_marx_names = [name.strip().lower() for name in app['arun_marx_info']['Name'].split(',')]

#                         for project_head in project_heads_names:
#                             if isinstance(project_head, str) and project_head.strip() != '':
#                                 project_head_lower = project_head.lower()
#                                 if any(name in project_head_lower for name in arun_marx_names):
#                                     project_lead_emails.append(app['arun_marx_info']['Email'])
#                                 else:
#                                     project_lead_email = df[df['Name'].str.lower() == project_head_lower]['Email Id'].unique()
#                                     project_lead_emails.extend(project_lead_email)
#                     project_lead_emails = list(set(project_lead_emails))

#                     st.write(f"To: {', '.join(map(str, group['Email Id'].unique()))}")
#                     st.write(f"CC: {', '.join(map(str, project_lead_emails))}")
#                     st.write("\nHello Team,")
#                     st.write(f"""
#                             {app['email_body']}
#                             \n{app['email_body_1']}
#                             \n{app['email_body_2']}
#                             \n{app['email_body_3']}
#                             \n***Here is the schedule for the week {start_of_week.strftime('%Y-%b-%d')} to {end_of_week.strftime('%Y-%b-%d')} for {client_mapping}***
#                             """)
#                     data = pd.DataFrame(group.drop(['Email Id'], axis=1).rename(columns=lambda x: x.strftime('%d-%b-%Y') if isinstance(x, datetime) else x))
#                     header_style = '''
#                         <style>
#                             th{
#                                 background-color: #236418;
#                                 color: white;
#                                 font-weight: bold;
#                             }
#                         </style>
#                     '''
#                     st.markdown(header_style, unsafe_allow_html=True)
#                     st.markdown(data.style.hide(axis="index").to_html(), unsafe_allow_html=True)
#                     st.write("\t")
#                     st.write("""\n
#                                 Best Regards,
#                                 \nIdeassion    \n""")

#             else:
#                 st.write("No 'Client Mapping' column found in the concatenated schedule.")
#         else:
#             st.write("No relevant data found for the specified date range.")

# if __name__ == "__main__":
#     st.set_page_config(layout="wide")
#     main()



import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import logging
import streamlit as st
from openpyxl import load_workbook
import warnings
import requests
import urllib.parse
from urllib.parse import urlparse, parse_qs
from dotenv import load_dotenv
import os

global client_id, tenant_id, client_secret, redirect_uri, authorization_base_url, token_endpoint, scope, app, send_mail_endpoint

with open('config.json') as f:
    app = json.load(f)

# Load environment variables from .env file
load_dotenv()

warnings.simplefilter(action='ignore', category=UserWarning)

client_id = os.getenv("CLIENT_ID")
tenant_id = os.getenv("TENANT_ID")
client_secret = os.getenv("CLIENT_SECRET")
redirect_uri = 'http://localhost:8501'
authorization_base_url = os.getenv("AUTHORIZATION_BASE_URL")
token_endpoint = os.getenv("TOKEN_ENDPOINT")
scope = "openid profile email"
send_mail_endpoint = os.getenv("send_mail_endpoint")

# Set up logging
logging.basicConfig(level=logging.WARNING)

def get_sheet_names(file_path):
    wb = load_workbook(filename=file_path)
    return wb.sheetnames

def get_authorization_url():
    params = {
        'client_id': client_id,
        'response_type': 'code',
        'redirect_uri': redirect_uri,
        'response_mode': 'query',
        'scope': scope
    }
    return f'{authorization_base_url}?{urllib.parse.urlencode(params)}'

def get_token_from_code(code):
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    body = {
        'grant_type': 'authorization_code',
        'code': code,
        'redirect_uri': redirect_uri,
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': scope
    }
    response = requests.post(token_endpoint, headers=headers, data=body)
    return response.json()

def fetch_user_data(token):
    headers = {
        'Authorization': f'Bearer {token}'
    }
    response = requests.get('https://graph.microsoft.com/v1.0/me', headers=headers)
    return response.json()

def get_access_token():
    auth_url = get_authorization_url()
    st.write(f'Click [here]({auth_url}) to log in.')

    code = st.experimental_get_query_params().get('code')
    if code:
        code = code[0]
        token_response = get_token_from_code(code)
        if 'access_token' in token_response:
            access_token = token_response['access_token']
            expires_in = token_response.get('expires_in')
            expiration_time = datetime.now() + timedelta(seconds=expires_in)
            refresh_token = token_response.get('refresh_token')
            return access_token, expiration_time, refresh_token
        else:
            st.error('Authentication failed.')
            return None, None, None

def refresh_access_token(refresh_token):
    token_request_params = {
        'client_id': client_id,
        'client_secret': client_secret,
        'refresh_token': refresh_token,
        'grant_type': 'refresh_token',
        'scope': scope
    }
    token_response = requests.post(token_endpoint, data=token_request_params)
    token_data = token_response.json()
    expires_in = token_data.get('expires_in')
    expiration_time = datetime.now() + timedelta(seconds=expires_in)
    return token_data.get('access_token'), expiration_time

def mail_auth(access_token, From, to, cc, Subject, email_body):
    headers = {
        'Authorization': 'Bearer ' + access_token,
        'Content-Type': 'application/json'
    }
    to_recipients = to.split(',')

    to_recipient_list = [{"emailAddress": {"address": email.strip()}} for email in to_recipients]
    cc_recipients = []
    if cc and not pd.isna(cc):
        cc_recipients_list = cc.split(',')
        cc_recipients = [{"emailAddress": {"address": email.strip()}} for email in cc_recipients_list]
    
    request_body = {
        "message": {
            "subject": f"{Subject}",
            "body": {
                "contentType": "HTML",
                "content": f"{email_body}"
            },
            "toRecipients": to_recipient_list,
            "ccRecipients": cc_recipients
        },
        "saveToSentItems": "true"
    }
    response = requests.post(send_mail_endpoint, headers=headers, json=request_body)
    if response.status_code == 202:
        return 'Email sent successfully!'
    else:
        return f'Email sending failed with status code: {response.status_code}'

def send_email(access_token, receiver_emails, cc_emails, group, start_date, end_date, client_mapping):
    sender_email = 'your_email@example.com'
    subject = f"Weekly Schedule - {start_date.strftime('%Y-%b-%d')} to {end_date.strftime('%Y-%b-%d')} - Client Mapping: {client_mapping}"

    From = sender_email
    Subject = subject

    header_style = '''
    <style>
        th {
            background-color: #236418;
            color: white;
            font-weight: bold;
        }
    </style>
    '''

    email_body = f"""<p>Hello Team,</p>
                    <p><b>{app['email_body']}</b></p>
                    <p><b>{app['email_body_1']}</b></p>
                    <p><b>{app['email_body_2']}</b></p>
                    <p><b>{app['email_body_3']}</b></p>                    
                    <p><b>Here is the schedule for the week {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')} for Client Mapping: {client_mapping}<b></p>
                    <p><b>Important: This is an Automated mail, if any mistakes or doubts please, approach HR</b></p>
                    {header_style}
                    {group.drop(['Project head', 'Email Id', 'Client Mapping'], axis=1).rename(columns=lambda x: x.strftime('%Y-%b-%d') if isinstance(x, datetime) else x).to_html(index=False)}
                    <p><b>Best regards,<br>Ideassion</b></p>
                    """
    
    try:
        mail_auth(access_token, From, receiver_emails, cc_emails, Subject, email_body)
        print("Email sent successfully.")
    except Exception as e:
        logging.warning(f"An error occurred while sending the email: {str(e)}")
        raise Exception(f"An error occurred while sending the email: {str(e)}")

def main():
    st.image('ticjdbcnfxi1xrkyrkol.png', width=300)
    st.title("Automatic Weekly Scheduler")
    try:
        today = datetime.today()
        start_of_week = today + timedelta(days=(7 - today.weekday()))
        end_of_week = start_of_week + timedelta(days= 6)
        st.write(f"Automatically scheduling for the week {start_of_week.strftime('%Y-%b-%d')} to {end_of_week.strftime('%Y-%b-%d')}.")
    except Exception as e:
        logging.warning('The date series not available')

    file_path = st.file_uploader("Upload Excel file", type=["xlsx"])
    if file_path is not None:
        try:
            df = pd.read_excel(file_path)
        except Exception as e:
            logging.error(f"An error occurred while reading the Excel file: {str(e)}")
            return

        desired_sheet_names = get_sheet_names(file_path)
        concatenated_schedules = []
        multiple_sheets = []

        for i, sheet_name in enumerate(desired_sheet_names):
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=False)
                df['Emp.ID'] = df['Emp.ID'].fillna(0).astype(int)
                for column in df.columns:
                    if column != 'Emp.ID':
                        df[column] = df[column].map(lambda x: '' if isinstance(x, int) else x)

                date_columns = [col for col in df.columns if isinstance(col, datetime) and
                                col.date() >= start_of_week.date() and col.date() <= end_of_week.date() and col.weekday() not in [5, 6]]

                if date_columns:
                    multiple_sheets.append(i)

                    relevant_columns = ['Emp.ID','Name', 'Project head', 'Email Id', 'Client Mapping']
                    if i == 0:
                        date_schedule = df[relevant_columns + date_columns]
                    elif i + 1:
                        if i - 1 in multiple_sheets:
                            date_schedule = df[date_columns]
                        else:
                            date_schedule = df[relevant_columns + date_columns]

                    concatenated_schedules.append(date_schedule)

                    current_sheet_data = df[relevant_columns + date_columns]

                    if concatenated_schedules:
                        concatenated_schedules[0].loc[:, 'Name'] = concatenated_schedules[0]['Name'].combine_first(
                            current_sheet_data['Name'])
                        concatenated_schedules[0].loc[:, 'Emp.ID'] = concatenated_schedules[0]['Emp.ID'].combine_first(
                            current_sheet_data['Emp.ID'])
                        concatenated_schedules[0].loc[:, 'Project head'] = concatenated_schedules[0]['Project head'].combine_first(
                            current_sheet_data['Project head'])
                        concatenated_schedules[0].loc[:, 'Email Id'] = concatenated_schedules[0]['Email Id'].combine_first(
                            current_sheet_data['Email Id'])
                        concatenated_schedules[0].loc[:, 'Client Mapping'] = concatenated_schedules[0]['Client Mapping'].combine_first(
                            current_sheet_data['Client Mapping'])
                    else:
                        concatenated_schedules.append(current_sheet_data)
            except Exception as e:
                logging.warning(f"An error occurred while processing sheet '{sheet_name}': {str(e)}")
                st.write(f"An error occurred while processing sheet '{sheet_name}': {str(e)}")

        if multiple_sheets:
            concatenated_schedule = pd.concat(concatenated_schedules, axis=1)

        if concatenated_schedules:
            concatenated_schedule = pd.concat(concatenated_schedules, axis=1)

            if 'Client Mapping' in concatenated_schedule.columns:
                grouped_schedule = concatenated_schedule.groupby('Client Mapping', as_index=False)      
                send_email_button = st.button("Send Email")
                if send_email_button:
                    access_token, expiration_time, refresh_token = get_access_token()

                    if datetime.now() > expiration_time:
                        print("Token has been expired")
                        access_token, expiration_time = refresh_access_token(refresh_token)

                    for client_mapping, group in grouped_schedule:
                        group['Project head'] = group['Project head'].replace('nan', np.NaN)
                        group = group.replace(np.nan, '', regex=True)

                        project_lead_emails = []
                        if 'arun_marx_info' in app:
                            arun_marx_names = [name.strip().lower() for name in app['arun_marx_info']['Name'].split(',')]
                            project_heads_names = set([name.strip() for head in group['Project head'] for name in head.split(',')])
                            for project_head in project_heads_names:
                                if isinstance(project_head, str) and project_head.strip() != '':
                                    project_head_lower = project_head.lower()
                                    if any(name in project_head_lower for name in arun_marx_names):
                                        project_lead_emails.append(app['arun_marx_info']['Email'])
                                    else:
                                        project_lead_email = df[df['Name'].str.lower() == project_head_lower]['Email Id'].unique()
                                        project_lead_emails.extend(project_lead_email)
                                project_lead_emails = list(set(project_lead_emails))
                        to = ', '.join(map(str, group['Email Id'].unique()))
                        cc = ', '.join(map(str, project_lead_emails))
                        send_email(access_token, to, cc, group, start_of_week, end_of_week, client_mapping)
                        st.write(f"Email sent for Client Mapping: {client_mapping}")

                for client_mapping, group in grouped_schedule:
                    st.subheader(f"Client Mapping: {client_mapping}")
                    group['Project head'] = group['Project head'].replace('nan', np.NaN)
                    group = group.replace(np.nan, '', regex=True)

                    project_heads_names = set([name.strip() for head in group['Project head'] for name in head.split(',')])

                    project_lead_emails = []
                    if 'arun_marx_info' in app:
                        arun_marx_names = [name.strip().lower() for name in app['arun_marx_info']['Name'].split(',')]

                        for project_head in project_heads_names:
                            if isinstance(project_head, str) and project_head.strip() != '':
                                project_head_lower = project_head.lower()
                                if any(name in project_head_lower for name in arun_marx_names):
                                    project_lead_emails.append(app['arun_marx_info']['Email'])
                                else:
                                    project_lead_email = df[df['Name'].str.lower() == project_head_lower]['Email Id'].unique()
                                    project_lead_emails.extend(project_lead_email)
                    project_lead_emails = list(set(project_lead_emails))

                    st.write(f"To: {', '.join(map(str, group['Email Id'].unique()))}")
                    st.write(f"CC: {', '.join(map(str, project_lead_emails))}")
                    st.write("\nHello Team,")
                    st.write(f"""
                            {app['email_body']}
                            \n{app['email_body_1']}
                            \n{app['email_body_2']}
                            \n{app['email_body_3']}
                            \n***Here is the schedule for the week {start_of_week.strftime('%Y-%b-%d')} to {end_of_week.strftime('%Y-%b-%d')} for {client_mapping}***
                            """)
                    data = pd.DataFrame(group.drop(['Email Id'], axis=1).rename(columns=lambda x: x.strftime('%d-%b-%Y') if isinstance(x, datetime) else x))
                    header_style = '''
                        <style>
                            th{
                                background-color: #236418;
                                color: white;
                                font-weight: bold;
                            }
                        </style>
                    '''
                    st.markdown(header_style, unsafe_allow_html=True)
                    st.markdown(data.style.hide(axis="index").to_html(), unsafe_allow_html=True)
                    st.write("\t")
                    st.write("""\n
                                Best Regards,
                                \nIdeassion    \n""")

            else:
                st.write("No 'Client Mapping' column found in the concatenated schedule.")
        else:
            st.write("No relevant data found for the specified date range.")

if __name__ == "__main__":
    st.set_page_config(layout="wide")
    main()
